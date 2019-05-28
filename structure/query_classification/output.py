import openpyxl
import os
from walle.settings import BASE_DIR
from pprint import pprint

from adwords_reports.report import AdwordsReport
from adwords_entities.batchjob import BatchJob


class Output(object):

	def __init__(self):
		self.client = None
		self.queriesInput = None
		self.keywordOps = []
		self.agOps = []
		self.campaignOps = []
		self.adOps = []
		self.budgetOps = []
		self.campaigns = None


	def setClient(self, client):
		self.client = client

	def setQueries(self, list_queries):
		self.queriesInput = list_queries
		if self.client:
			self.getQueryAds()

	def addQuery(self, query):
		self.queriesInput.append(query)

	def getQueryAds(self):
		ag_ids = [query.getAdGroupId() for query in self.queriesInput]
		
		query = ("SELECT AdGroupId, Id, HeadlinePart1, HeadlinePart2, Description, Path1, Path2, CreativeFinalUrls, Status " +
				"FROM AD_PERFORMANCE_REPORT " +
				"WHERE AdGroupId IN " + str(ag_ids))

		report = AdwordsReport(query, self.client).returnDataAsListOfDicts()

		ads = {}
		for ad in report:
			if ad["AdGroupId"] not in ads:
				ads[ad["AdGroupId"]] = []
			ads[ad["AdGroupId"]].append(ad)

		for query in self.queriesInput:
			query.setAds(ads[query.getAdGroupId()])


	def buildExcelOutput(self):
		
		output_keywords = [["Campaign", "Adgroup", "Keyword", "MatchType", "Final Url"]]
		output_ads = [["Campaign", "Adgroup", "Headline1", "Headline2", "Description", "Path1", "Path2", "FinalUrl"]]
		output_negatives = [["Campaign", "Adgroup", "Keyword", "MatchType"]]
		output_custom_params = [["Campaign", "Adgroup", "Custom Parameters"]]
		output_custom_params_camp = [["Campaign", "Custom Parameters"]]

		for query in self.queriesInput:
			if "negative" in query.getTags():
				continue

			campaign = query.getCampaignName()
			ag = query.getAdGroupName()
			keyword = query.getName()
			matchtype = query.getMatchTypeOutput()
			final_url = query.getFinalUrl()

			output_keywords.append([campaign, ag, keyword, matchtype, final_url])
			output_negatives.append([campaign, ag, keyword.replace("+", ""), "EXACT"])
			output_custom_params.append([campaign, ag, "{_group}="+ag])

			# CustomParam Campaign
			row_to_append = [campaign, "{_campaign}="+campaign]
			if row_to_append not in output_custom_params_camp:
				output_custom_params_camp.append(row_to_append)

			# Ads
			if query.getAds():
				for ad in query.getAds():
					hl1 = ad["HeadlinePart1"]
					hl2 = ad["HeadlinePart2"]
					description = ad["Description"]
					final_url = ad["CreativeFinalUrls"][4:][:-4]
					output_ads.append([campaign, ag, hl1, hl2, description, "", "", final_url])

		return {"keywords": output_keywords, "ads": output_ads, "negatives": output_negatives, "customParameters": output_custom_params, "customParametersCampaign": output_custom_params_camp}


	def outputToXlsx(self):
		"""	Input: Object of lists
			Output: String download file location
		"""
		# get 2d-lists from input object
		data = self.buildExcelOutput()

		entities = [	data["keywords"],
						data["ads"],
						data["negatives"],
						data["customParameters"],
						data["customParametersCampaign"]	]

		# create excel sheets in workbook
		wb = openpyxl.Workbook()
		sheets = [	wb.active,
					wb.create_sheet(title="Ads"),
					wb.create_sheet(title="Negatives"),
					wb.create_sheet(title="Custom Parameters"),
					wb.create_sheet(title="Custom Params Campaign")	]
		
		# fill sheets
		for index, entity in enumerate(entities):
			for row in entity:
				sheets[index].append(row)

		# save to desired destination
		dest_filename = os.path.join(BASE_DIR, "structure/media/output.xlsx")
		wb.save(filename=dest_filename)

		return "structure/media/output.xlsx"

	""" ///////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
		API FUNCTIONS
		///////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
	"""

	def getCampaigns(self):
		query = ("SELECT CampaignId, CampaignName, CampaignStatus, BudgetId, Amount, BudgetId, TrackingUrlTemplate " +
				"FROM CAMPAIGN_PERFORMANCE_REPORT ")

		campaigns = AdwordsReport(query, self.client).returnDataAsListOfDicts()	
		campaigns_by_name = {}

		for campaign in campaigns:
			camp_name = campaign["CampaignName"]
			if camp_name not in campaigns_by_name:
				campaigns_by_name[camp_name] = campaign

		self.campaigns = campaigns_by_name



	def sendToApi(self):
		bj = BatchJob(self.client)
		self.getCampaigns()

		for query in self.queriesInput:
			#deal with campaign id
			camp_name = query.getCampaignName()

			if camp_name in self.campaigns:
				camp_id = self.campaigns[camp_name]["CampaignId"]
			else:
				camp_id = bj.getHelperId()
				budget_id = bj.getHelperId()
				# setting standard budget to 10 EUR/USD
				bj.addBudgetOps(budget_id, camp_name, 10000000)
				bj.addCampaignOps(camp_id, camp_name, budget_id)

			bj_ag_id = bj.getHelperId()
			bj.addAdGroupOps(camp_id, bj_ag_id, query.getAdGroupName(), query.getAdGroupBid())
			bj.addKeywordOps(bj_ag_id, query.getName(), query.getMatchTypeOutput(), query.getFinalUrl())
			bj.addNegativeOps(bj_ag_id, query.getName(), query.getMatchTypeOutput())
			bj.addAdOps(bj_ag_id, query.getAds())

		pprint(bj.campaignOps)
		pprint(bj.agOps)
		pprint(bj.kwOps)
